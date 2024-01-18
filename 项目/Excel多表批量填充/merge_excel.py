import openpyxl

def read_data_from_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data_array = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  
        if any(cell is not None for cell in row):
            data_array.append(row)
    workbook.close()
    return data_array

def write_data_to_excel(data_array, output_file, sheet_name):
    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook[sheet_name]
    # print(sheet.max_row)

    for row_data in data_array:
        sheet.append(row_data)

    workbook.save(output_file)

def main():
    # 读取第一个 Excel 文件中 Sheet2 的数据
    input_file = './test.xlsx'
    input_sheet = '模板'
    data_array = read_data_from_excel(input_file, input_sheet)
    print(data_array)
    # 将数据写入第二个 Excel 文件中 Sheet2，保留表头
    output_file = './test2.xlsx'
    output_sheet = '模板'
    write_data_to_excel(data_array, output_file, output_sheet)

if __name__ == "__main__":
    main()
